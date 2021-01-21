#!/usr/bin/env python3
# coding=utf-8

import os
import openpyxl
import allure
import traceback
import time
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from loguru import logger

from common.data_util import read_excel_sheet
from page.base_page import BasePage
from common.excel import ParseExcel
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from common.file_util import get_work_path, get_curr_time_str
from common import data_util


class MdesBaseCase(BasePage):
    """
    每个系统有各自的数据读取方法，目前这个方法自定义
    """

    """
    ==================================================================
    对EXCEL文件的操作（读取，创建，写入）
    ==================================================================
    """

    @staticmethod
    def load_data(data_file):
        """ 创建测试数据
        从excel读取测试数据拼接为字典 """
        data_list = []  # 创建一个列表存放测试数据
        excel = ParseExcel(data_file)

        """ 初始化读取列起点 """
        sheet1_num, sheet2_num, sheet3_num = (0, 0, 0)

        """"""
        (sheet1_name,sheet2_name,sheet3_name) = ('投保事项', '客户信息', '告知及授权')

        """"""
        sheet1 = excel.get_sheet_by_name(sheet1_name)
        sheet2 = excel.get_sheet_by_name(sheet2_name)
        sheet3 = excel.get_sheet_by_name(sheet3_name)
        """ 获得每张表的最大列 """
        sheet1_max = (excel.get_cols_num(sheet1) - 1)
        sheet2_max = (excel.get_cols_num(sheet2) - 1)
        sheet3_max = (excel.get_cols_num(sheet3) - 1)

        """ 比较各表最大列数取得循环结束点 """
        max_cycle = max(sheet1_max, sheet2_max, sheet3_max)
        i = 1
        for i in range(max_cycle - 1):
            sheet1_num += 1
            if sheet1_num >= sheet1_max:
                sheet1_num = 1
            sheet2_num += 1
            if sheet2_num >= sheet2_max:
                sheet2_num = 1
            sheet3_num += 1
            if sheet3_num >= sheet3_max:
                sheet3_num = 1
            i += 1

            a = data_util.handle_data(read_excel_sheet('datas/read_mdes.xlsx', sheet1_name, [sheet1_num])[0])
            b = data_util.handle_data(read_excel_sheet('datas/read_mdes.xlsx', sheet2_name, [sheet2_num])[0])
            c = data_util.handle_data(read_excel_sheet('datas/read_mdes.xlsx', sheet3_name, [sheet3_num])[0])

            temp_dict = {**a, **b, **c}
            data_list.append(temp_dict)
        return data_list


    @staticmethod
    def write_dictinlist_into_sheet(result_list):
        """
        将列表中字典的内容，循环写到excel文件里.
        result_list：列表字典[{},{}...]
        """
        print('看看写了几遍？？？')
        relative_paths_file = 'datas\\mdes_' + get_curr_time_str() + '.xlsx'
        file_path = os.path.join(get_work_path(), relative_paths_file)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '五要素和险种信息'
        ws.cell(1, 1, 'key')
        key_cols = []
        for i, item in enumerate(result_list):  # 循环取出数组i下标，item值
            ws.cell(column=i + 2, row=1).value = 'value' + str(i + 1)  # 用i来控制列
            for key, value in item.items():  # 取出列表中以索引下标的各个字典，key和value循环
                if key not in key_cols:  # 如果k不在 当前的key列表中，则进行添加
                    key_cols.append(key)  # 然后再将这个新的key追加到key的列表中
                ws.cell(column=1, row=(key_cols.index(key) + 2)).value = key  # 将key添加到第1列 key所在列表位置的行中(excel 1起始+1行(key,value1..标题)
                ws.cell(column=i + 2, row=(key_cols.index(key) + 2)).value = value # 将最大的列和判断这个行数是通过这个key所在的位置+2得到，+2（标题行算1）
            wb.save(file_path)
            wb.close()


    """
    ==================================================================
    定位操作
    ==================================================================
    """

    def switch_frame_to_main(self):
        """ 从frame切换回根目录 """
        try:
            self.driver.switch_to.default_content()
        except Exception as e:
            print(traceback.print_exc())

    def get_elements(self, locator, sec=60):
        """
        """
        by, value, comment = self._handle_locator(locator)
        elements=self.driver.find_elements(by, value)
        return elements

    def select_option_native(self, locator, v):
        """ html原生select标签 """
        by, value, comment = self._handle_locator(locator)
        step_info = "使用{0}={1}方式查找下拉框{2},并根据文本值{3}来选择".format(by, value, comment, v)
        with allure.step(step_info):
            element = Select(self.find_element(locator))
            element.select_by_visible_text(v)
            logger.info((step_info))

    def input_text_value(self, locator, value):
        """ html原生输入(修改value) """
        element = self.find_element(locator)
        self.driver.execute_script("(arguments[0]).value=arguments[1];", element, value)

    def wait_until_element_not_visible(self, locator, time_out=360):
        by, value, comment = self._handle_locator(locator)
        WebDriverWait(self.driver, time_out, 1).until_not(lambda x: x.find_element(by=by, value=value).is_displayed(),
                                                          message="等待元素方式{0}={1}不显示未成功".format(by, value))

    def popup_confirm(self):
        """ 弹出确认/删除窗口
         点击确认 """
        box = self.driver.switch_to_alert()
        time.sleep(2)
        print(box.text)
        box.accept()

    def click_element_by_js(self, locator, timeout=15):
        """
        使用js模拟点击操作
        """
        by, value, comment = self._handle_locator(locator)
        with allure.step("使用{0}={1}方式查询元素{2},并进行点击".format(by, value, comment)):
            # element = self.find_element(locator, timeout=timeout)
            element = self.find_element(locator)
            self.driver.execute_script("arguments[0].click();", element)

    def get_element_text(self, locator):
        """ 获取元素文本值 """
        value = self.find_element(locator).text
        return value

    def get_url(self, ):
        url_add = self.driver.get('http://10.164.26.48:9081/mdes/banklns/banklnsFrame.jsf?type=2')
        return url_add

    """
    ==================================================================
    等待操作
    ==================================================================
    """

    def wait_element(self, locator, sec=15):
        """
        等待元素出现
        :param locator: 定位方法+定位表达式组合字符串，用逗号分隔，如'css,.username'
        :param sec:等待秒数
        """
        by, value, commit = self._handle_locator(locator)
        try:
            WebDriverWait(self.driver, sec, 1).until(lambda x: x.find_element(by=by, value=value),
                                                     message='element not found!!!')
            return True
        except TimeoutException:
            return False
        except Exception as e:
            raise e

    def wait_until_page_contains(self, text, sec=15):
        """ 等待文字出现 """
        locator = (By.XPATH, '//*[contains(.,"%s")]' % text)
        by, value = self._handle_locator(locator)
        try:
            WebDriverWait(self.driver, sec, 1).until(lambda x: x.find_element(by=by, value=value),
                                                     message='text not found!!!')
            return True
        except TimeoutException:
            return False
        except Exception as e:
            raise e

    # def wait_until_element_not_visible(self, locator, time_out=360):
    #     by, value, comment = self._handle_locator(locator)
    #     WebDriverWait(self.driver, time_out, 1).until_not(lambda x: x.find_element(by=by, value=value).is_displayed(),
    #                                                       message="等待元素方式{0}={1}不显示未成功".format(by, value))


if __name__ == '__main__':

    mdesbasecase = MdesBaseCase(webdriver)
    # result_list =[{'appnt_ID': 'S3144143312', 'appnt_ID_type': '03', 'appnt_birthday': '19800201', 'appnt_name': '理赔难'},{'appnt_ID': 'S3144143312', 'appnt_ID_type': '03', 'appnt_birthday': '19800201', 'appnt_name': '理赔难','aaa': '111','bbb': '222'}]
    # mdesbasecase.write_dictinlist_into_sheet(result_list)

    a = MdesBaseCase.load_data('datas/read_mdes.xlsx')
    print((a))

