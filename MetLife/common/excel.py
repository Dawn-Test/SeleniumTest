from openpyxl import load_workbook
import os
import allure
import re
from common.file_util import get_work_path
from common.parser import parse_data, extract_variables, extract_functions
from httprunner.loader.load import load_module_functions
import importlib


class ParseExcel(object):

    def __init__(self, filename):
        file_path = os.path.join(get_work_path(), filename)
        self.wb = load_workbook(file_path)
        self.excelFile = file_path

    def get_sheet_by_name(self, sheet_name):
        """获取sheet对象"""
        sheet = self.wb[sheet_name]
        return sheet

    def get_row_num(self, sheet):
        """获取有效数据的最大行号"""
        return sheet.max_row

    def get_cols_num(self, sheet):
        """获取有效数据的最大列号"""
        return sheet.max_column

    def get_row_values(self, sheet, row_num):
        """获取某一行的数据"""
        max_cols_num = self.get_cols_num(sheet)
        row_values = []
        for colsNum in range(1, max_cols_num + 1):
            value = sheet.cell(row_num, colsNum).value
            if value is None:
                value = ''
            row_values.append(value)
        return tuple(row_values)

    def get_column_values(self, sheet, column_num):
        """获取某一列的数据"""
        max_row_num = self.get_row_num(sheet)
        column_values = []
        for rowNum in range(2, max_row_num + 1):
            value = sheet.cell(rowNum, column_num).value
            if value is None:
                value = ''
            column_values.append(value)
        return tuple(column_values)

    def get_value_of_cell(self, sheet, row_num, column_num):
        """获取某一个单元格的数据"""
        value = sheet.cell(row_num, column_num).value
        if value is None:
            value = ''
        return value

    @allure.step('获取测试数据')
    def get_all_values_of_sheet(self, sheet):
        """获取某一个sheet页的所有测试数据，返回一个元祖组成的列表"""
        with allure.step("从工作表{0}取得测试数据，返回为字典".format(sheet)):
            max_row_num = self.get_row_num(sheet)
            column_num = self.get_cols_num(sheet)
            all_values = []
            for row in range(2, max_row_num + 1):
                row_values = []
                for column in range(3, column_num + 1):
                    value = sheet.cell(row, column).value
                    if value is None:
                        value = ''
                    elif value.startswith('$'):
                        value = self.generate_data_by_expr(value)
                    row_values.append(value)
                all_values.append(row_values)
            data = dict(all_values)
            return data

    def generate_data_by_expr(self, data):
        """ 根据数据公式创建对应测试数据 """
        if isinstance(data, str) and data.startswith('$'):
            comp = re.compile('[^A-Z^a-z^0-9^_]')
            method = comp.sub('', data)
            class_name = 'DataTools'
            module_meta = __import__('common.data_tools', globals(), locals(), [class_name])
            class_meta = getattr(module_meta, class_name)
            cls_obj = class_meta()
            """ hasattr判断方法是否存在 """
            print(hasattr(cls_obj, method))
            if hasattr(cls_obj, method):
                func = getattr(cls_obj, method)
                data_value = func()
                return data_value
            return data

    def get_value_of_key(self, sheet, key):
        """ 根据key查找value """
        max_row_num = self.get_row_num(sheet)
        for rowNum in range(2, max_row_num + 1):
            keyvalue = sheet.cell(rowNum, 2).value
            if keyvalue == key:
                a = rowNum
                break
        value = self.get_value_of_cell(sheet, a, 3)
        return value

    def write_into_sheet(self, sheet, key, value):
        """ 将页面数据写入表格
        追加写入，不覆盖既有数据，key为数据注释，value为是数据值 """
        max_row_num = self.get_row_num(sheet)
        max_row_num += 1
        print(max_row_num)
        sheet['A' + str(max_row_num)] = key
        sheet['B' + str(max_row_num)] = value
        self.wb.save(self.excelFile)
        self.wb.close()


    def get_case_values_of_sheet(self, sheet, case_no):
        """ 从表格里获取case对应的测试数据，返回一个字典 """
        with allure.step("从工作表{0}取得case{1}测试数据，返回为字典".format(sheet, case_no)):
            max_row_num = self.get_row_num(sheet)
            # print(max_row_num)
            excel_key_value = {}
            for row in range(2, max_row_num+1):
                data_key = sheet.cell(row, 2).value
                data_value = sheet.cell(row, 2 + case_no).value
                if data_value is None or data_value == '':
                    data_value = ''
                excel_key_value[data_key] = data_value
        variables_mapping = {}
        # imported_module = importlib.import_module("common.functions.built_in_functions")
        imported_module = importlib.import_module("common.functions.built_in_functions")
        functions_mapping = load_module_functions(imported_module)

        imported_module1 = importlib.import_module("common.functions.mdes_buile_in_functions")
        aaa = load_module_functions(imported_module1)
        functions_mapping.update(aaa)
        while True:
            excel_key_value = parse_data(excel_key_value, variables_mapping=variables_mapping,
                                         functions_mapping=functions_mapping,
                                         raise_if_variable_not_found=False)
            count = 0
            for k, v in excel_key_value.items():
                ev = extract_variables(v)
                ef = extract_functions(v)
                if ev != [] or ef != []:
                    count += 1
            if count == 0:
                break
        return excel_key_value


if __name__ == '__main__':
    excel = ParseExcel('datas\\read_mdes.xlsx')
    sheet = excel.get_sheet_by_name('客户信息')
    a = excel.get_case_values_of_sheet(sheet, 1)
    print(a)
