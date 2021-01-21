#!/usr/bin/env python3
# coding=utf-8
import importlib

from httprunner.loader.load import load_module_functions
from openpyxl import load_workbook

from common.file_util import get_work_path
from common.parser import *

VARIABLE_COLUMN = 1
DATA_START_ROW = 2
DATA_START_COLUMN = 3


def read_excel_sheet(excel_file, sheet_name, select_ids=[]):
    """
    读取指定excel文件中的内容，并以列表形式返回
    excel_file ：excel文件，输入为相对路径
    sheet_name : sheet_name
    select_ids ：支持选取第几条数据
    """
    data_file_path = os.path.join(get_work_path(), excel_file)
    wb = load_workbook(data_file_path)
    ws = wb[sheet_name]

    if select_ids:
        select_ids = [int(i) + DATA_START_COLUMN - 1 for i in select_ids]
    else:
        select_ids = range(DATA_START_COLUMN, ws.max_column + 1)
    test_case_data_list = []
    for i in select_ids:
        test_case_data = {}
        for j in range(DATA_START_ROW, ws.max_row + 1):  #(2,55+1)
            variable_name = ws.cell(row=j, column=VARIABLE_COLUMN).value  # 从第二行第一列开始取key
            variable_value = ws.cell(row=j, column=i).value  #(2,3) # 从第二行第三列开始取value
            test_case_data[variable_name] = variable_value  #生成个字典
        test_case_data_list.append(test_case_data)  #再追加到列表中
    return test_case_data_list


def hand_functions_map():
    module_list = get_functions_modules()
    functions_mapping = {}
    for name in module_list:
        imported_module = importlib.import_module(name)
        custom_functions_map = load_module_functions(imported_module)
        functions_mapping.update(custom_functions_map)
    return functions_mapping


def handle_data(content, variables_map=None, functions_map=None):
    variables_map = variables_map if variables_map else {}
    built_in_functions_mapping = hand_functions_map()
    functions_map = functions_map.update(
        built_in_functions_mapping) if functions_map else built_in_functions_mapping
    while True:
        content = parse_data(content, variables_mapping=variables_map, functions_mapping=functions_map,
                             raise_if_variable_not_found=False)
        count = 0
        for k, v in content.items():
            ev = extract_variables(v)
            ef = extract_functions(v)
            if ev != [] or ef != []:
                count += 1
        if count == 0:
            break
    return content


def get_functions_modules():
    """
    获取包名下所有非__init__的模块名
    """
    modules = []
    functions_path = os.path.join(os.path.dirname(__file__), 'functions')
    files = os.listdir(functions_path)

    for file in files:
        if not file.startswith("__") and file.endswith('.py'):
            name, ext = os.path.splitext(file)
            modules.append("common.functions." + name)
    return modules


if __name__ == '__main__':
    data_list =read_excel_sheet('datas/read_mdes.xlsx', '投保事项',[1,2])
    print(type(data_list))


    a = handle_data(read_excel_sheet('datas/read_mdes.xlsx', '投保事项')[0])
    b = handle_data(read_excel_sheet('datas/read_mdes.xlsx', '投保事项')[1])
    print(a)
    print(b)




